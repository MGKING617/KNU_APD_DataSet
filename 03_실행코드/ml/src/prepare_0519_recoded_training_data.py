from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd


DEFAULT_INPUT = "data/raw/0519) 지역사회건강조사 재범주화_수정본.xlsx"
DEFAULT_DATA_SHEET = "0519) 지역사회건강조사 재범주화"
DEFAULT_CUTOFF_SHEET = "cut-off 기준"
DEFAULT_OUTPUT = "data/processed/chs_training_0519_recoded.csv"
DEFAULT_REPORT_DIR = "ml/reports/0519_recode"
PHQ_SOURCE_COLUMNS = [f"mtb_07{suffix}1" for suffix in "abcdefghi"]
TARGET_COLUMN = "mh_PHQ_S"


FEATURE_MAP = {
    "age": "age",
    "mbhld_co": "household_size",
    "주관적건강기준": "subjective_health",
    "흡연경험여부": "smoking_status",
    "연간음주빈도": "alcohol_frequency",
    "걷기실천일수": "walking_days",
    "체중조절경험여부": "weight_control",
    "고혈압진단여부": "hypertension_diagnosis",
    "당뇨병진단여부": "diabetes_diagnosis",
    "연간사고중독여부": "accident_poisoning_experience",
    "기초생활수급자여부": "basic_livelihood_recipient",
    "아침식사일수": "breakfast_frequency",
    "경제활동여부": "economic_activity",
    "성별": "sex",
    "거주지": "region_group",
    "교육수준": "education_level",
    "혼인상태": "marital_status",
    "주관적스트레스": "stress_level",
}

FEATURE_COLUMNS = [
    "age",
    "household_size",
    "subjective_health",
    "smoking_status",
    "alcohol_frequency",
    "walking_days",
    "body_self_thin",
    "body_self_overweight",
    "weight_control",
    "hypertension_diagnosis",
    "diabetes_diagnosis",
    "accident_poisoning_experience",
    "basic_livelihood_recipient",
    "breakfast_frequency",
    "economic_activity",
    "sex",
    "region_group",
    "education_level",
    "marital_status",
    "stress_level",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare the 0519 recoded CHS training dataset.")
    parser.add_argument("--input", default=DEFAULT_INPUT)
    parser.add_argument("--data-sheet", default=DEFAULT_DATA_SHEET)
    parser.add_argument("--cutoff-sheet", default=DEFAULT_CUTOFF_SHEET)
    parser.add_argument("--output", default=DEFAULT_OUTPUT)
    parser.add_argument("--report-dir", default=DEFAULT_REPORT_DIR)
    parser.add_argument("--min-age", type=int, default=19)
    parser.add_argument("--max-age", type=int, default=39)
    parser.add_argument("--threshold", type=float, default=10.0)
    return parser.parse_args()


def numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def derive_phq_score(raw: pd.DataFrame) -> pd.Series:
    missing = [column for column in PHQ_SOURCE_COLUMNS if column not in raw.columns]
    if missing:
        raise ValueError(f"Missing PHQ source columns: {missing}")
    phq = raw[PHQ_SOURCE_COLUMNS].apply(pd.to_numeric, errors="coerce")
    valid = phq.where(phq.isin([1, 2, 3, 4]))
    return (valid - 1).sum(axis=1, min_count=len(PHQ_SOURCE_COLUMNS))


def add_body_features(processed: pd.DataFrame, raw: pd.DataFrame) -> None:
    if "본인인지체형" not in raw.columns:
        raise ValueError("Missing body image recoded column: 본인인지체형")
    body = numeric(raw["본인인지체형"])
    processed["body_self_thin"] = np.where(body.eq(0), 1.0, np.where(body.isna(), np.nan, 0.0))
    processed["body_self_overweight"] = np.where(body.eq(2), 1.0, np.where(body.isna(), np.nan, 0.0))


def prepare_dataset(raw: pd.DataFrame, min_age: int, max_age: int) -> tuple[pd.DataFrame, dict[str, Any]]:
    missing = [column for column in list(FEATURE_MAP) + ["본인인지체형"] if column not in raw.columns]
    if missing:
        raise ValueError(f"Missing source columns: {missing}")

    processed = pd.DataFrame(index=raw.index)
    for source, target in FEATURE_MAP.items():
        processed[target] = numeric(raw[source])
    add_body_features(processed, raw)
    processed[TARGET_COLUMN] = derive_phq_score(raw)

    before_rows = len(processed)
    target_valid_rows = int(processed[TARGET_COLUMN].notna().sum())
    age_valid = processed["age"].between(min_age, max_age)
    output = processed[age_valid & processed[TARGET_COLUMN].notna()].copy()
    output = output[FEATURE_COLUMNS + [TARGET_COLUMN]]

    metadata = {
        "source_rows": int(before_rows),
        "target_valid_rows": target_valid_rows,
        "output_rows": int(len(output)),
        "dropped_rows": int(before_rows - len(output)),
        "dropped_rows_reason": {
            "invalid_or_missing_phq_score": int(processed[TARGET_COLUMN].isna().sum()),
            "outside_age_range_or_missing_age": int((~age_valid).sum()),
        },
        "age_filter": [min_age, max_age],
        "target_column": TARGET_COLUMN,
        "target_formula": "sum(mtb_07a1..mtb_07i1 - 1) when all nine source values are 1..4",
        "feature_columns": FEATURE_COLUMNS,
        "excluded_columns": PHQ_SOURCE_COLUMNS + ["본인인지체형"],
        "exclusion_policy": {
            "phq_items": "Target calculation only, not model features.",
            "본인인지체형": "Replaced by body_self_thin and body_self_overweight binary features.",
        },
    }
    return output, metadata


def text_color(part: Any) -> str:
    if type(part).__name__ != "TextBlock":
        return "plain"
    font = getattr(part, "font", None)
    color = getattr(font, "color", None)
    rgb = str(getattr(color, "rgb", "")).upper()
    if rgb == "FFFF0000":
        return "red"
    if rgb == "FF0070C0":
        return "blue"
    return "black"


def cell_parts(cell: Any) -> list[dict[str, str]]:
    value = cell.value
    if type(value).__name__ == "CellRichText":
        return [
            {
                "text": str(getattr(part, "text", part)).replace("\n", " / "),
                "color": text_color(part),
            }
            for part in value
        ]
    return [{"text": "" if value is None else str(value).replace("\n", " / "), "color": "plain"}]


def values_from_text(text: str) -> list[int]:
    values: list[int] = []
    for start, end in re.findall(r"(?<!\d)(\d+)(?:\s*~\s*(\d+)\s*일?)?", text):
        if end:
            values.extend(range(int(start), int(end) + 1))
        else:
            values.append(int(start))
    return values


def build_cutoff_audit(input_path: Path, cutoff_sheet: str) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(input_path, data_only=True, read_only=False, rich_text=True)
    worksheet = workbook[cutoff_sheet]
    records: list[dict[str, Any]] = []
    for row_index in range(1, worksheet.max_row + 1):
        variable = worksheet.cell(row_index, 1).value
        if not variable:
            continue
        source_parts = cell_parts(worksheet.cell(row_index, 3))
        recoded_parts = cell_parts(worksheet.cell(row_index, 4))
        grouped_values: dict[str, list[int]] = {"red": [], "blue": [], "black": [], "plain": []}
        for part in source_parts:
            grouped_values.setdefault(part["color"], []).extend(values_from_text(part["text"]))
        recoded_values = sorted(
            {int(match) for match in re.findall(r"(?<!\d)(\d+)\s*=", str(worksheet.cell(row_index, 4).value or ""))}
        )
        records.append(
            {
                "source_sheet": cutoff_sheet,
                "source_row": row_index,
                "variable": str(variable),
                "display_name": str(worksheet.cell(row_index, 2).value or ""),
                "source_text": str(worksheet.cell(row_index, 3).value or "").replace("\n", " / "),
                "recoded_text": str(worksheet.cell(row_index, 4).value or "").replace("\n", " / "),
                "red_source_values": ";".join(map(str, sorted(set(grouped_values["red"])))),
                "blue_source_values": ";".join(map(str, sorted(set(grouped_values["blue"])))),
                "black_or_missing_source_values": ";".join(map(str, sorted(set(grouped_values["black"] + grouped_values["plain"])))),
                "recoded_values": ";".join(map(str, recoded_values)),
                "is_binary_recoded": recoded_values == [0, 1],
            }
        )
    return pd.DataFrame(records)


def value_distribution(frame: pd.DataFrame, columns: list[str]) -> dict[str, dict[str, int]]:
    output: dict[str, dict[str, int]] = {}
    for column in columns:
        counts = frame[column].value_counts(dropna=False).sort_index().to_dict()
        output[column] = {str(key): int(value) for key, value in counts.items()}
    return output


def validation_report(
    raw: pd.DataFrame,
    processed: pd.DataFrame,
    metadata: dict[str, Any],
    threshold: float,
    input_path: Path,
    data_sheet: str,
    cutoff_sheet: str,
) -> dict[str, Any]:
    target = (processed[TARGET_COLUMN] >= threshold).astype(int)
    missing_rates = processed[FEATURE_COLUMNS].isna().mean().sort_values(ascending=False)
    single_value_columns = [
        column
        for column in FEATURE_COLUMNS
        if processed[column].dropna().nunique() <= 1
    ]
    non_binary_features = []
    for column in FEATURE_COLUMNS:
        values = set(processed[column].dropna().unique().tolist())
        if column not in {"age", "household_size"} and not values.issubset({0, 1, 0.0, 1.0}):
            non_binary_features.append({"column": column, "values": sorted(map(float, values))})
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "input_file": str(input_path),
        "data_sheet": data_sheet,
        "cutoff_sheet": cutoff_sheet,
        "metadata": metadata,
        "target": {
            "threshold": threshold,
            "positive_rows": int(target.sum()),
            "negative_rows": int((1 - target).sum()),
            "positive_rate": float(target.mean()),
            "score_min": float(processed[TARGET_COLUMN].min()),
            "score_max": float(processed[TARGET_COLUMN].max()),
            "score_mean": float(processed[TARGET_COLUMN].mean()),
        },
        "raw_columns": list(raw.columns),
        "feature_columns": FEATURE_COLUMNS,
        "missing_rate_by_feature": {column: float(value) for column, value in missing_rates.items()},
        "value_distribution_by_feature": value_distribution(processed, FEATURE_COLUMNS + [TARGET_COLUMN]),
        "non_binary_features_except_continuous": non_binary_features,
        "single_value_feature_columns": single_value_columns,
        "raw_value_distribution": value_distribution(raw, list(raw.columns)),
    }


def write_mapping_table(report_dir: Path) -> None:
    rows = [
        {
            "source_column": source,
            "processed_column": target,
            "note": "Copied from the 0519 recoded workbook.",
        }
        for source, target in FEATURE_MAP.items()
    ]
    rows.extend(
        [
            {
                "source_column": "본인인지체형",
                "processed_column": "body_self_thin",
                "note": "1 when recoded body group is 0=thin; normal and overweight are 0; missing stays missing.",
            },
            {
                "source_column": "본인인지체형",
                "processed_column": "body_self_overweight",
                "note": "1 when recoded body group is 2=overweight; thin and normal are 0; missing stays missing.",
            },
            {
                "source_column": ",".join(PHQ_SOURCE_COLUMNS),
                "processed_column": TARGET_COLUMN,
                "note": "PHQ-like score from source values 1..4 mapped to 0..3 and summed; excluded from features.",
            },
        ]
    )
    pd.DataFrame(rows).to_csv(report_dir / "feature_mapping_0519.csv", index=False, encoding="utf-8-sig")


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    report_dir = Path(args.report_dir)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_dir.mkdir(parents=True, exist_ok=True)

    raw = pd.read_excel(input_path, sheet_name=args.data_sheet)
    processed, metadata = prepare_dataset(raw, args.min_age, args.max_age)
    processed.to_csv(output_path, index=False, encoding="utf-8-sig")

    cutoff_audit = build_cutoff_audit(input_path, args.cutoff_sheet)
    cutoff_audit.to_csv(report_dir / "cutoff_mapping_audit_0519.csv", index=False, encoding="utf-8-sig")
    write_mapping_table(report_dir)

    report = validation_report(
        raw,
        processed,
        metadata,
        args.threshold,
        input_path,
        args.data_sheet,
        args.cutoff_sheet,
    )
    (report_dir / "validation_report_0519.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    output_path.with_suffix(".summary.json").write_text(
        json.dumps(
            {
                "rows": int(len(processed)),
                "positive_rows": report["target"]["positive_rows"],
                "negative_rows": report["target"]["negative_rows"],
                "positive_rate": report["target"]["positive_rate"],
                "columns": list(processed.columns),
                "feature_columns": FEATURE_COLUMNS,
                "source_file": str(input_path),
                "report_dir": str(report_dir),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(
        json.dumps(
            {
                "output": str(output_path),
                "report_dir": str(report_dir),
                "rows": int(len(processed)),
                "positive_rows": report["target"]["positive_rows"],
                "negative_rows": report["target"]["negative_rows"],
                "features": len(FEATURE_COLUMNS),
                "non_binary_features_except_continuous": report["non_binary_features_except_continuous"],
                "single_value_feature_columns": report["single_value_feature_columns"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
